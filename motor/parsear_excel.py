#!/usr/bin/env python3
"""
parsear_excel.py — Lee un Excel de participante y produce
datos/pronosticos/<porra>/<nickname>.json

Uso:
    python motor/parsear_excel.py <ruta_excel> <porra>
    porra: 'amigos' o 'trabajo'
"""

import json
import re
import sys
import unicodedata
import warnings
from pathlib import Path

import openpyxl

# ── Rutas ────────────────────────────────────────────────────────────────────
BASE = Path(__file__).resolve().parent.parent
CALENDARIO    = BASE / "datos" / "calendario.json"
EQUIVALENCIAS = BASE / "datos" / "equivalencias_equipos.json"
PRONOSTICOS   = BASE / "datos" / "pronosticos"


# ── Normalización ────────────────────────────────────────────────────────────
def norm(texto: str | None) -> str:
    """Minúsculas sin tildes ni espacios extra. Útil para comparaciones fuzzy."""
    if not texto:
        return ""
    t = unicodedata.normalize("NFD", str(texto))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return t.strip().lower()


# ── Carga de referencias ─────────────────────────────────────────────────────
def _cargar_calendario() -> list[dict]:
    with open(CALENDARIO, encoding="utf-8") as f:
        return json.load(f)["partidos"]


def _cargar_equivalencias() -> list[dict]:
    with open(EQUIVALENCIAS, encoding="utf-8") as f:
        return json.load(f)["selecciones"]


def _build_nombre_index(equivalencias: list[dict]) -> dict[str, str]:
    """
    Devuelve {forma_normalizada -> nombre_openfootball}.
    Cubre nombre_excel, nombre_openfootball, nombre_api_football y aliases_es.
    """
    idx: dict[str, str] = {}
    for sel in equivalencias:
        canonical = sel["nombre_openfootball"]
        for campo in ("nombre_excel", "nombre_openfootball", "nombre_api_football"):
            idx[norm(sel[campo])] = canonical
        for alias in sel.get("aliases_es", []):
            idx[norm(alias)] = canonical
    return idx


def _build_match_index(partidos: list[dict]) -> dict[tuple, int]:
    """
    Índice para partidos de grupos: (grupo, jornada, norm(local)) -> match_id
    """
    return {
        (p["grupo"], p["jornada"], norm(p["local"])): p["id"]
        for p in partidos
        if p["fase"] == "grupos"
    }


# ── Parseo de código de jornada ───────────────────────────────────────────────
_RE_JORNADA = re.compile(r"^([A-L])([123])$")

def _parsear_jornada(codigo) -> tuple[str | None, str | None]:
    """'A1' -> ('A', 'J1')  |  'L3' -> ('L', 'J3')  |  None -> (None, None)"""
    if not codigo:
        return None, None
    m = _RE_JORNADA.match(str(codigo).strip())
    if not m:
        return None, None
    return m.group(1), f"J{m.group(2)}"


# ── Parseo de predicción de grupos ───────────────────────────────────────────
_RE_PRED = re.compile(r"^([1X2])\|(\d+)-(\d+)$")

# ── Posiciones de grupo (filas 80-127, col B label + col C equipo) ────────────
_RE_GROUP_POS = re.compile(r"^(1st|2nd|3rd|4th)\s+GROUP\s+([A-L])$", re.IGNORECASE)
_ORD_MAP = {"1st": 1, "2nd": 2, "3rd": 3, "4th": 4}

# ── Marcadores de eliminatoria (col C: "Local-Visitante·SIGNO|GL-GV") ─────────
_RE_ELIM_PRED = re.compile(r"·([1X2])\|(\d+)-(\d+)$")

def _parsear_pred_grupo(valor) -> dict | None:
    """
    '1|2-0' -> {signo:'1', goles_local:2, goles_visitante:0}
    'X|1-1' -> {signo:'X', goles_local:1, goles_visitante:1}
    '0|-' o None o vacío -> None (sin pronóstico)
    """
    if not valor:
        return None
    m = _RE_PRED.match(str(valor).strip())
    if not m:
        return None
    return {
        "signo": m.group(1),
        "goles_local": int(m.group(2)),
        "goles_visitante": int(m.group(3)),
    }


# ── Detección de placeholders de la plantilla ────────────────────────────────
# Cubre: 2A, 1C, 3ABCDF, W73, L101, WF, LF, W34, W97-W98·0|-, etc.
_RE_PLACEHOLDER = re.compile(
    r"^("
    r"[12][A-L]"           # 2A, 1C …
    r"|3[A-L]{2,}"         # 3ABCDF, 3CDFGH …
    r"|W\d+"               # W73, W97 …
    r"|L\d+"               # L101, L102 …
    r"|WF|LF|W34"          # honor placeholders
    r"|[A-L][1-4]"         # A1, B2 (residuales)
    r"|.*·.*"              # match predictions con '·' (bloques de resultados)
    r")$"
)

def _es_placeholder(valor) -> bool:
    if not valor:
        return True
    return bool(_RE_PLACEHOLDER.match(str(valor).strip()))


def _resolver_equipo(valor, nombre_idx: dict[str, str]) -> str | None:
    """Convierte nombre del Excel al nombre canónico (openfootball). None si placeholder."""
    if _es_placeholder(valor):
        return None
    clave = norm(str(valor))
    return nombre_idx.get(clave)  # None si no reconocido


def _split_match_elim(b_val, nombre_idx: dict[str, str]) -> tuple[str | None, str | None]:
    """
    Divide "LocalTeam-AwayTeam" (col B) en nombres canónicos.
    Prueba cada '-' como separador para manejar nombres con guion.
    """
    s = str(b_val).strip()
    pos = 0
    while True:
        idx = s.find("-", pos)
        if idx == -1:
            break
        cand_l = s[:idx].strip()
        cand_v = s[idx + 1:].strip()
        l_can = nombre_idx.get(norm(cand_l))
        v_can = nombre_idx.get(norm(cand_v))
        if l_can and v_can:
            return l_can, v_can
        pos = idx + 1
    # Fallback: primer guion
    parts = s.split("-", 1)
    if len(parts) == 2:
        return (
            nombre_idx.get(norm(parts[0].strip()), parts[0].strip()),
            nombre_idx.get(norm(parts[1].strip()), parts[1].strip()),
        )
    return s, None


def _parsear_posiciones_grupo(ws, nombre_idx: dict[str, str]) -> list[dict]:
    """
    Lee filas 80-127 de la hoja Pool.
    Col B: etiqueta "1st GROUP A" … "4th GROUP L"
    Col C: nombre del equipo
    Devuelve 48 entradas (12 grupos × 4 posiciones).
    No existe bloque separado de mejores terceros en el Excel; las posiciones 3
    de cada grupo ya están incluidas aquí con pos=3.
    """
    posiciones: list[dict] = []
    for fila in range(80, 128):
        etiqueta  = ws.cell(fila, 2).value  # col B
        equipo_raw = ws.cell(fila, 3).value  # col C
        if not etiqueta or not equipo_raw:
            continue
        m = _RE_GROUP_POS.match(str(etiqueta).strip())
        if not m:
            continue
        pos    = _ORD_MAP[m.group(1).lower()]
        grupo  = m.group(2).upper()
        equipo = nombre_idx.get(norm(str(equipo_raw)), str(equipo_raw).strip())
        entrada: dict = {"grupo": grupo, "pos": pos, "equipo": equipo}
        if pos == 3:
            entrada["mejor_tercero"] = True
        posiciones.append(entrada)
    return posiciones


def _parsear_elim_marcadores(ws, nombre_idx: dict[str, str]) -> list[dict]:
    """
    Lee los marcadores de los partidos de eliminatoria desde col C.
    Formato celda: "Local-Visitante·SIGNO|GL-GV"
    """
    RONDAS: list[tuple[str, list[int]]] = [
        ("1/16",  list(range(164, 180))),
        ("1/8",   list(range(200, 208))),
        ("1/4",   list(range(220, 224))),
        ("semis", list(range(232, 234))),
        ("3-4",   [244]),
        ("final", [247]),
    ]
    marcadores: list[dict] = []
    for ronda, filas in RONDAS:
        for fila in filas:
            col_b = ws.cell(fila, 2).value
            col_c = ws.cell(fila, 3).value
            if not col_b or not col_c:
                continue
            m = _RE_ELIM_PRED.search(str(col_c).strip())
            if not m:
                continue
            local, visitante = _split_match_elim(col_b, nombre_idx)
            marcadores.append({
                "ronda":    ronda,
                "local":    local,
                "visitante": visitante,
                "signo":    m.group(1),
                "gl":       int(m.group(2)),
                "gv":       int(m.group(3)),
            })
    return marcadores


# ── Parser principal ──────────────────────────────────────────────────────────
def parsear_excel(ruta_excel: Path, porra: str) -> dict:
    partidos     = _cargar_calendario()
    equivalencias = _cargar_equivalencias()
    nombre_idx   = _build_nombre_index(equivalencias)
    match_idx    = _build_match_index(partidos)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)

    ws   = wb["Pool"]
    ws_p = wb["Premios"]

    # ── Nickname ──────────────────────────────────────────────────────────────
    raw_nick = ws["C5"].value
    nickname = str(raw_nick).strip() if raw_nick else ""
    if nickname.lower() in ("", "nombre"):
        nickname = None

    # ── Partidos de grupos (filas 6–77) ───────────────────────────────────────
    pronosticos_grupos: list[dict] = []
    for fila in range(6, 78):
        cod     = ws[f"A{fila}"].value
        col_b   = ws[f"B{fila}"].value
        col_c   = ws[f"C{fila}"].value

        grupo, jornada = _parsear_jornada(cod)
        if not grupo or not col_b:
            continue

        # Local = parte antes del primer '-' en columna B
        partes     = str(col_b).split("-", 1)
        local_exc  = partes[0].strip()
        local_can  = nombre_idx.get(norm(local_exc), local_exc)

        match_id = match_idx.get((grupo, jornada, norm(local_can)))

        pronosticos_grupos.append({
            "match_id":   match_id,
            "grupo":      grupo,
            "jornada":    jornada,
            "local":      local_can,
            "visitante":  nombre_idx.get(norm(partes[1].strip())) if len(partes) > 1 else None,
            "prediccion": _parsear_pred_grupo(col_c),
        })

    # ── Bloques de eliminatorias ───────────────────────────────────────────────
    def _leer_bloque(filas) -> list[str | None]:
        return [_resolver_equipo(ws[f"C{r}"].value, nombre_idx) for r in filas]

    teams_1_16   = _leer_bloque(range(130, 162))   # 32 equipos
    teams_1_8    = _leer_bloque(range(182, 198))   # 16 equipos
    teams_1_4    = _leer_bloque(range(210, 218))   #  8 equipos
    teams_semis  = _leer_bloque(range(226, 230))   #  4 equipos
    teams_3_4    = _leer_bloque(range(236, 238))   #  2 (partido 3º-4º)
    teams_final  = _leer_bloque(range(240, 242))   #  2 finalistas

    # ── Honor (C250-C252) ─────────────────────────────────────────────────────
    campeon    = _resolver_equipo(ws["C250"].value, nombre_idx)
    subcampeon = _resolver_equipo(ws["C251"].value, nombre_idx)
    tercero    = _resolver_equipo(ws["C252"].value, nombre_idx)

    # Derivar 4º: el equipo del partido 3º-4º que no sea el 3º
    cuarto = None
    if tercero and any(teams_3_4):
        for eq in teams_3_4:
            if eq and eq != tercero:
                cuarto = eq
                break
    # Fallback: buscar en semis
    if cuarto is None:
        finales_set = {e for e in teams_final if e}
        for eq in teams_semis:
            if eq and eq not in finales_set and eq != tercero:
                cuarto = eq
                break

    # ── Posiciones de grupo ───────────────────────────────────────────────────
    posiciones_grupo = _parsear_posiciones_grupo(ws, nombre_idx)

    # ── Marcadores de eliminatoria ────────────────────────────────────────────
    elim_marcadores = _parsear_elim_marcadores(ws, nombre_idx)

    # ── Premios de jugador ────────────────────────────────────────────────────
    def _texto(v) -> str | None:
        if v is None:
            return None
        s = str(v).strip()
        return s or None

    goleador = _texto(ws_p["B4"].value)
    mvp      = _texto(ws_p["B5"].value)
    portero  = _texto(ws_p["B6"].value)

    # ── Comprobación de completitud ───────────────────────────────────────────
    advertencias: list[str] = []

    if not nickname:
        advertencias.append("nickname no rellenado (C5)")

    n_sin_pred = sum(1 for p in pronosticos_grupos if p["prediccion"] is None)
    if n_sin_pred:
        advertencias.append(f"{n_sin_pred}/72 partidos de grupos sin pronóstico")

    n_none_1_16 = sum(1 for e in teams_1_16 if e is None)
    if n_none_1_16:
        advertencias.append(f"{n_none_1_16}/32 equipos de 1/16 sin rellenar")

    if not all([goleador, mvp, portero]):
        faltantes = [k for k, v in [("goleador", goleador), ("MVP", mvp), ("portero", portero)] if not v]
        advertencias.append(f"premios incompletos: faltan {', '.join(faltantes)}")

    resultado = {
        "nickname":        nickname,
        "porra":           porra,
        "archivo_origen":  ruta_excel.name,
        "incompleto":      bool(advertencias),
        "advertencias":    advertencias,
        "pronosticos": {
            "grupos": pronosticos_grupos,
            "posiciones_grupo": posiciones_grupo,
            "elim_marcadores": elim_marcadores,
            "clasificados": {
                "1/16":  teams_1_16,
                "1/8":   teams_1_8,
                "1/4":   teams_1_4,
                "semis": teams_semis,
                "final": teams_final,
            },
            "honor": {
                "campeon":    campeon,
                "subcampeon": subcampeon,
                "tercero":    tercero,
                "cuarto":     cuarto,
            },
            "premios": {
                "goleador": goleador,
                "mvp":      mvp,
                "portero":  portero,
            },
        },
    }
    return resultado


# ── Punto de entrada ──────────────────────────────────────────────────────────
def main() -> None:
    if len(sys.argv) < 3:
        print("Uso: python motor/parsear_excel.py <ruta_excel> <porra>")
        print("     porra: 'amigos' o 'trabajo'")
        sys.exit(1)

    ruta   = Path(sys.argv[1])
    porra  = sys.argv[2].lower()

    if porra not in ("amigos", "trabajo"):
        print(f"ERROR: porra debe ser 'amigos' o 'trabajo', recibido: '{porra}'")
        sys.exit(1)
    if not ruta.exists():
        print(f"ERROR: no se encuentra el archivo {ruta}")
        sys.exit(1)

    resultado = parsear_excel(ruta, porra)

    nick_safe = re.sub(r"[^\w\-]", "_", resultado["nickname"] or "sin_nickname")
    destino   = PRONOSTICOS / porra / f"{nick_safe}.json"
    destino.parent.mkdir(parents=True, exist_ok=True)

    with open(destino, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    estado = "INCOMPLETO" if resultado["incompleto"] else "COMPLETO"
    print(f"[{estado}]  {ruta.name}  ->  {destino}")
    for adv in resultado["advertencias"]:
        print(f"  !  {adv}")


if __name__ == "__main__":
    main()
